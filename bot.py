import pyautogui
import pyperclip
from time import sleep
import webbrowser
import PySimpleGUI as sg  # package PySimpleGui
import pandas as pd  # package pandas
from openpyxl import load_workbook
df1 = pd.DataFrame()


def menu():  # Janela 1
    sg.theme('Dark Blue 3')
    layout = [[sg.Text('Bem-vindo(a) ao RPA Bitrix v2.0 !!\n'
                       'O que deseja fazer?:\n')],
              [sg.Button('RPA'), sg.Button('Cancelar')],
              [sg.Text('\nIndicium Tech - 2022', size=[75, 5], justification='center')]]
    return sg.Window('Menu', layout=layout, finalize=True, size=(500, 180))


def erro():  # Janela 2
    sg.theme('DarkRed')
    layout = [[sg.Text('Favor verificar:\n'
                       '\n1) Planilha deve estar fechada\n'
                       '2) Extensão do arquivo de saída estar em formato .xlsx\n'
                       '3) Dados na planilha fora do padrão\n'
                       '4) Não há dados a ser processado\n')],
              [sg.Button('Voltar'), sg.Button('Cancelar')]]
    return sg.Window('ERRO', layout=layout, size=(500, 200), finalize=True)


def sucesso():  # Janela 3
    sg.theme('DarkGreen')
    layout = [[sg.Text('Processo realizado com sucesso !')],
              [sg.Button('Voltar'), sg.Button('Cancelar')]]
    return sg.Window('SUCESSO', layout=layout, size=(300, 100), finalize=True)


def rpa():  # Janela 5
    sg.theme('Dark Blue 3')
    layout = [
              [sg.Text('Caminho do Arquivo do Google Sheets')],
              [sg.Input(), sg.FileBrowse(key='-SAIDA-', file_types=(('Text Files', '*.xls'),
                                                                    ('Text Files', '*.xlsx')))],
              [sg.Button('OK'), sg.Button('Voltar'), sg.Button('Cancelar')],
             [sg.Text('\nIndicium Tech - 2021', size=[75, 5], justification='center')]]
    return sg.Window('RPA - Lançamento no Bitrix', layout=layout, finalize=True, size=(600, 180))


janela1, janela2, janela3, janela4, janela5 = menu(), None, None, None, None

while True:
    window, event, values = sg.read_all_windows()
    # Operações no MENU
    if window == janela1 and event == sg.WINDOW_CLOSED:
        break
    if window == janela1 and event == 'Cancelar':
        break
    if window == janela1 and event == 'RPA':
        janela1.close()
        janela5 = rpa()
    if window == janela2 and event == 'Voltar':
        janela2.close()
    if window == janela2 and event == 'Cancelar':
        break
    if window == janela2 and event == sg.WINDOW_CLOSED:
        break
    if window == janela3 and event == 'Voltar':
        janela3.close()
        janela5.close()
        janela1 = menu()
    if window == janela3 and event == 'Cancelar':
        break
    if window == janela3 and event == sg.WINDOW_CLOSED:
        break
    if window == janela5 and event == 'OK' and values['-SAIDA-'] != '':
        try:
            path_saida = values['-SAIDA-']
            df1 = pd.DataFrame(pd.read_excel(path_saida, sheet_name='Tarefas'))
            webbrowser.open('https://google.com.br')
            sleep(2)
            for i, row in df1.iterrows():
                id = (str(row['ID']))
                campoCTI = (str(row['CTI']))
                rpa = (str(row['RPA']))
                if id != 'nan' and campoCTI != 'nan' and rpa != 'Processado':
                    webbrowser.open('https://indicium.bitrix24.com/workgroups/group/' + str(row['Cod projeto']) +
                                    '/tasks/task/edit/' + str(row['ID']) + '/')
                    sleep(20)
                    pyperclip.copy('')  # Teste
                    pyautogui.hotkey('ctrl', 'a'), sleep(1)  # Teste
                    pyautogui.hotkey('ctrl', 'c'), sleep(1)
                    descricao = pyperclip.paste()
                    print(descricao)
                    if descricao != '':
                        df1.loc[i, 'Descricao'] = descricao
                        book = load_workbook(path_saida)
                        writer = pd.ExcelWriter(path_saida, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df1.to_excel(writer, "Tarefas", header=True, index=False)
                        writer.save()
                    try:
                        pyautogui.hotkey('ctrl', 'f'), sleep(1)
                        pyperclip.copy('Campo CTI')
                        pyautogui.hotkey('ctrl', 'v'), sleep(1)
                        pyautogui.hotkey('enter'), sleep(1)
                        pyautogui.hotkey('esc'), sleep(1)
                        pyautogui.hotkey('tab'), sleep(1)
                        pyperclip.copy(str(row['CTI']))
                        pyautogui.hotkey('ctrl', 'v'), sleep(1)
                        pyautogui.hotkey('ctrl', 'f'), sleep(1)
                        pyperclip.copy('Salvar Alterações')
                        pyautogui.hotkey('ctrl', 'v'), sleep(1)
                        pyautogui.hotkey('enter'), sleep(1)
                        pyautogui.hotkey('esc'), sleep(1)
                        pyautogui.hotkey('enter'), sleep(1)
                        sleep(15)
                        pyautogui.hotkey('ctrl', 'w')
                        df1.loc[i, 'RPA'] = 'Processado'
                        book = load_workbook(path_saida)
                        writer = pd.ExcelWriter(path_saida, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df1.to_excel(writer, "Tarefas", header=True, index=False)
                        writer.save()
                    except:
                        sleep(10)
                        pyautogui.hotkey('ctrl', 'w')
                        df1.loc[i, 'RPA'] = 'ERRO'
                        book = load_workbook(path_saida)
                        writer = pd.ExcelWriter(path_saida, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df1.to_excel(writer, "Tarefas", header=True, index=False)
                        writer.save()
                else:
                    df1.loc[i, 'RPA'] = 'Faltam dados para ser possível processar'
                    book = load_workbook(path_saida)
                    writer = pd.ExcelWriter(path_saida, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df1.to_excel(writer, "Tarefas", header=True, index=False)
                    writer.save()
            janela1.close()
            janela3 = sucesso()
        except:
            janela1.close()
            janela2 = erro()
    if window == janela5 and event == 'OK' and values['-SAIDA-'] == '':
        janela1.close()
        janela2 = erro()
    if window == janela5 and event == 'Voltar':
        janela5.close()
        janela1 = menu()
    if window == janela5 and event == 'Cancelar':
        break
    if window == janela5 and event == sg.WINDOW_CLOSED:
        break
